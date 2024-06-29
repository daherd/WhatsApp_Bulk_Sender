from openpyxl import Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep




def run_webdriver(excel_file_path, massage):
    wb = load_workbook(excel_file_path)
    ws = wb.active

    text = massage
    web = webdriver.Chrome()
    flag=1
    for x in range(1, ws.max_row + 1):
        phone = ws.cell(row=x, column=1).value
        phone = str(phone)
        phone = phone.replace("-", "")
        if len(phone) >= 10:
            # phone=phone[:10]
            print(x, phone, len(phone))
            link = ('https://wa.me/' + '972' + phone + '?text=' + text)
            web.get(link)
            web.implicitly_wait(3)
            try:
                web.find_element(By.XPATH,
                                 '/html/body/div[1]/div[1]/div[2]/div/section/div/div/div/div[2]/div[1]/a/span').click()
            except:
                web.implicitly_wait(1)
                continue
            web.implicitly_wait(3)
            web.find_element(By.XPATH,
                             '/html/body/div[1]/div[1]/div[2]/div/section/div/div/div/div[3]/div/div/h4[2]/a/span').click()
            if flag == 1:
                flag=0
                web.implicitly_wait(40)
            else:
                web.implicitly_wait(3)


            try:
                sleep(4)
                web.find_element(By.XPATH,
                                 '/html/body/div[1]/div/div[2]/div[4]/div/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span').click()
            except:
                web.implicitly_wait(1)
            sleep(3)
    sleep(10)
    web.close()


run_webdriver("Book1.xlsx","%D8%A7%D9%84%D8%B3%D9%84%D8%A7%D9%85+%D8%B9%D9%84%D9%8A%D9%83%D9%85+%D8%A7%D8%AE%D9%8A+%D8%A7%D9%84%D9%83%D8%B1%D9%8A%D9%85%0D%0A%D9%86%D8%B0%D9%83%D8%B1%D9%83+%D8%A8%D8%B4%D8%AF+%D8%A7%D9%84%D8%B1%D8%AD%D8%A7%D9%84+%D8%A7%D9%84%D9%89+%D8%A7%D9%84%D9%85%D8%B3%D8%AC%D8%AF+%D8%A7%D9%84%D8%A3%D9%82%D8%B5%D9%89+%D8%BA%D8%AF%D8%A7+%D8%A5%D9%86+%D8%B4%D8%A7%D8%A1+%D8%A7%D9%84%D9%84%D9%87.%0D%0A%D8%A7%D9%84%D8%A7%D9%86%D8%B7%D9%84%D8%A7%D9%82+%D9%85%D9%86+%D8%A7%D9%85%D8%A7%D9%85+%D8%A8%D9%88%D8%A7%D8%A8%D8%A9+5+%D8%A7%D9%84%D8%B3%D8%A7%D8%B9%D8%A9+15%3A00.%0D%0A%D8%A7%D9%84%D8%A8%D8%B1%D9%86%D8%A7%D9%85%D8%AC+%D9%8A%D8%B4%D9%85%D9%84+%D8%A7%D9%81%D8%B7%D8%A7%D8%B1+%D8%AC%D9%85%D8%A7%D8%B9%D9%8A+%D9%88%D8%B5%D9%84%D8%A7%D8%A9+%D8%AA%D8%B1%D8%A7%D9%88%D9%8A%D8%AD%0D%0A%D8%A7%D9%84%D8%B1%D8%AC%D8%A7%D8%A1+%D8%A7%D9%84%D8%A7%D9%84%D8%AA%D8%B2%D8%A7%D9%85+%D8%A8%D8%A7%D9%84%D9%88%D9%82%D8%AA%0D%0A%D9%86%D8%B1%D8%A7%D9%83+%D8%B9%D9%84%D9%89+%D8%AE%D9%8A%D8%B1+%0D%0A%0D%0A%D8%AA%D8%A3%D9%83%D9%8A%D8%AF+%D8%A7%D9%84%D8%AD%D8%B6%D9%88%D8%B1+%D9%85%D9%86+%D8%AE%D9%84%D8%A7%D9%84+%D8%A7%D9%84%D8%B1%D8%AF+%D8%B9%D9%84%D9%89+%D8%A7%D9%84%D8%B1%D8%B3%D8%A7%D9%84%D8%A9.")
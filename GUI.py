import tkinter as tk
from tkinter import filedialog
from functools import partial
import pandas as pd
#from whatsapp_sender import *
from openpyxl import Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from tkinter import PhotoImage

def urlencode(text):
    encoded_text = ""
    special_chars = {'!': '%21', '"': '%22', '#': '%23', '$': '%24', '&': '%26',
                     "'": '%27', '(': '%28', ')': '%29', '*': '%2A', '+': '%2B',
                     ',': '%2C', '/': '%2F', ':': '%3A', ';': '%3B', '=': '%3D',
                     '?': '%3F', '@': '%40', '[': '%5B', '\\': '%5C', ']': '%5D',
                     '^': '%5E', '`': '%60', '{': '%7B', '|': '%7C', '}': '%7D',
                     ' ': '%20', '\n': '%0A'}
    for char in text:
        if char.isalnum() or char in ['-', '_', '.', '~']:
            encoded_text += char
        elif char in special_chars:
            encoded_text += special_chars[char]
        else:
            encoded_text += '%' + '{:02X}'.format(ord(char))
    return encoded_text

# Example usage:
#text_to_encode = "Hello World! How are you?\nI'm fine, thank you."
#encoded_text = urlencode(text_to_encode)
#print("Encoded Text:")
#print(encoded_text)




def run_webdriver(excel_file_path, massage):
    wb = load_workbook(excel_file_path)
    ws = wb.active

    encoded_text = urlencode(massage)
    web = webdriver.Chrome()
    flag=1
    for x in range(1, ws.max_row + 1):
        phone = ws.cell(row=x, column=1).value
        phone = str(phone)
        phone = phone.replace("-", "")
        if len(phone) >= 10:
            # phone=phone[:10]
            print(x, phone, len(phone))
            link = ('https://wa.me/' + '972' + phone + '?text=' + encoded_text)
            web.get(link)
            web.implicitly_wait(3)
            try:
                web.find_element(By.XPATH,
                                 '/html/body/div[1]/div[1]/div[2]/div/section/div/div/div/div[2]/div[1]/a/span').click()
            except:
                web.implicitly_wait(1)
                continue
            web.implicitly_wait(3)
            web.find_element(By.XPATH,
                             '/html/body/div[1]/div[1]/div[2]/div/section/div/div/div/div[3]/div/div/h4[2]/a/span').click()
            if flag == 1:
                flag=0
                web.implicitly_wait(40)
            else:
                web.implicitly_wait(3)

            try:
                sleep(4)
                web.find_element(By.XPATH,
                                 '/html/body/div[1]/div/div[2]/div[4]/div/footer/div[1]/div/span[2]/div/div[2]/div[2]').click()
            except:
                web.implicitly_wait(1)
            sleep(3)
    sleep(10)
    web.close()


def open_file_entry(entry_widget):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, file_path)

def submit_function(file_entry, message_entry):
    file_path = file_entry.get()
    message = message_entry.get("1.0", tk.END)
    print(message)
    run_webdriver(file_path, message)

def create_gui():
    root = tk.Tk()
    root.title("WhatsUp Bulk Sender")

    # File Selection
    file_label = tk.Label(root, text="Select Excel File:")
    file_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
    file_entry = tk.Entry(root, width=50)
    file_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)
    file_button = tk.Button(root, text="Browse", command=partial(open_file_entry, file_entry))
    file_button.grid(row=0, column=2, padx=10, pady=5)

    # Message Entry
    message_label = tk.Label(root, text="Enter Message:")
    message_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
    message_entry = tk.Text(root, width=50, height= 10)
    message_entry.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)


    # Submit Button
    submit_button = tk.Button(root, text="Submit", command=partial(submit_function, file_entry, message_entry))
    submit_button.grid(row=2, column=1, padx=10, pady=5)

    root.mainloop()

if __name__ == "__main__":
    create_gui()